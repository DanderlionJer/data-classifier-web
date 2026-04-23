from __future__ import annotations

import csv
from datetime import datetime, timezone
from io import BytesIO, StringIO

from openpyxl import Workbook

from app.models import ClassifyResponse

_SEP = "\uff1b"


def _field_table_rows(resp: ClassifyResponse) -> list[list[object]]:
    headers: list[str] = [
        "level",
        "categories",
        "database",
        "schema",
        "table",
        "column",
        "data_type",
        "comment",
        "tags",
        "tags_zh",
        "rationale",
        "matched_rule_ids",
        "ai_baseline_level",
        "ai_review_suggested",
        "ai_note",
    ]
    rows: list[list[object]] = [headers]
    for row in resp.fields:
        fld = row.field
        cats = _SEP.join(c.label_zh or c.id for c in row.categories)
        tags = ",".join(row.tags)
        tags_zh = _SEP.join(t for t in row.tags_zh if str(t).strip())
        rule_ids = ",".join(m.rule_id for m in row.matches)
        ai_base: str | int = ""
        ai_rev = ""
        ai_note = ""
        if row.ai:
            ai_base = row.ai.baseline_level
            ai_rev = "true" if row.ai.review_suggested else "false"
            ai_note = row.ai.note or ""
        rows.append(
            [
                row.level,
                cats,
                fld.database,
                fld.schema_name,
                fld.table,
                fld.column,
                fld.data_type,
                fld.comment,
                tags,
                tags_zh,
                row.rationale,
                rule_ids,
                ai_base,
                ai_rev,
                ai_note,
            ]
        )
    return rows


def classify_response_to_xlsx(resp: ClassifyResponse) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "fields"
    for r in _field_table_rows(resp):
        ws.append(r)

    wsum = wb.create_sheet("summary")
    wsum.append(["level", "count"])
    for k in sorted(resp.summary.keys(), key=lambda x: int(x), reverse=True):
        wsum.append([k, resp.summary[k]])
    wsum.append([])
    wsum.append(["category_id", "count"])
    for cid, cnt in sorted(
        resp.category_summary.items(), key=lambda x: x[1], reverse=True
    ):
        wsum.append([cid, cnt])

    wmeta = wb.create_sheet("meta")
    wmeta.append(["key", "value"])
    wmeta.append(["exported_at_utc", datetime.now(timezone.utc).isoformat()])
    wmeta.append(["country", resp.country or ""])
    wmeta.append(["ai_enhancement_applied", str(resp.ai_enhancement_applied).lower()])
    wmeta.append(["ai_model", resp.ai_model or ""])
    wmeta.append(["ai_provider", resp.ai_provider or ""])
    wmeta.append(["applied_frameworks", _SEP.join(resp.applied_frameworks)])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def classify_response_to_csv(resp: ClassifyResponse) -> bytes:
    buf = StringIO()
    writer = csv.writer(buf)
    for row in _field_table_rows(resp):
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")
